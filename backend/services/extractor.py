"""
Turns an uploaded skill-gap document into text the LLM can read.

All formats (PDF, Excel, CSV, plain text) are converted to a text content
block, since Groq / LLaMA models are text-only and cannot process binary
files natively.

PDFs are extracted via pdfplumber, which preserves tables and layout
reasonably well. Spreadsheets and CSVs are flattened to pipe-delimited text.
"""

import csv
import io

MAX_BYTES = 25 * 1024 * 1024          # 25 MB — well under the 32 MB request cap
MAX_SHEET_ROWS = 500                   # guard against someone uploading 100k rows
MAX_TEXT_CHARS = 200_000

PDF_TYPES = {"application/pdf"}
EXCEL_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",                                          # .xls
}
CSV_TYPES = {"text/csv", "application/csv"}
TEXT_TYPES = {"text/plain", "text/markdown"}


class UnsupportedDocument(Exception):
    pass


def _kind(filename: str, content_type: str | None) -> str:
    name = (filename or "").lower()
    ct = (content_type or "").lower()

    if name.endswith(".pdf") or ct in PDF_TYPES:
        return "pdf"
    if name.endswith((".xlsx", ".xlsm", ".xls")) or ct in EXCEL_TYPES:
        return "excel"
    if name.endswith(".csv") or ct in CSV_TYPES:
        return "csv"
    if name.endswith((".txt", ".md")) or ct in TEXT_TYPES:
        return "text"

    raise UnsupportedDocument(
        f"Unsupported file type: {filename or content_type!r}. "
        "Upload a PDF, Excel (.xlsx/.xls), CSV, or plain-text file."
    )


def _excel_to_text(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    chunks: list[str] = []

    for sheet in wb.worksheets:
        chunks.append(f"### Sheet: {sheet.title}")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= MAX_SHEET_ROWS:
                chunks.append(f"... (truncated at {MAX_SHEET_ROWS} rows)")
                break
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                chunks.append(" | ".join(cells))
        chunks.append("")

    wb.close()
    return "\n".join(chunks)


def _csv_to_text(data: bytes) -> str:
    text = data.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    lines = [
        " | ".join(cell.strip() for cell in row)
        for row in rows[:MAX_SHEET_ROWS]
        if any(cell.strip() for cell in row)
    ]
    if len(rows) > MAX_SHEET_ROWS:
        lines.append(f"... (truncated at {MAX_SHEET_ROWS} rows)")
    return "\n".join(lines)


def _pdf_to_text(data: bytes) -> str:
    from pypdf import PdfReader

    chunks: list[str] = []
    reader = PdfReader(io.BytesIO(data))
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            chunks.append(f"### Page {i + 1}")
            chunks.append(text.strip())
            chunks.append("")
    return "\n".join(chunks)


def to_content_block(filename: str, content_type: str | None, data: bytes) -> dict:
    """
    Return a single text content block representing the uploaded document.

    All formats are converted to text since Groq / LLaMA models are text-only.
    """
    if not data:
        raise UnsupportedDocument("The uploaded file is empty.")
    if len(data) > MAX_BYTES:
        raise UnsupportedDocument(
            f"File is {len(data) // (1024 * 1024)} MB; the limit is {MAX_BYTES // (1024 * 1024)} MB."
        )

    kind = _kind(filename, content_type)

    if kind == "pdf":
        body = _pdf_to_text(data)
    elif kind == "excel":
        body = _excel_to_text(data)
    elif kind == "csv":
        body = _csv_to_text(data)
    else:
        body = data.decode("utf-8-sig", errors="replace")

    body = body.strip()
    if not body:
        raise UnsupportedDocument("No readable content found in the uploaded file.")

    return {
        "type": "text",
        "text": f"<uploaded_document filename=\"{filename}\">\n{body[:MAX_TEXT_CHARS]}\n</uploaded_document>",
    }
